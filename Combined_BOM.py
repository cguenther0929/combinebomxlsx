"""
FILE: Combined_BOM.py

PURPOSE: 
When running this script, the user will be able to create one large flat BOM
for purchasing purposes.  This script will only operate on .xls files, and not
.xlsx files.  This script will automatically sift through all files in the 
current working directory, and with each file, it will iterate over all sheets. 

AUTHOR: 
Clinton G. 

TODO: Nothing

"""
import sys
import random
import time
import csv
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd		
import xlwt

## DEFINE VRIABLES ##
#####################
MFGPN_col 	= 0						# Column number containing the MFGPN
QPN_col 	= 0						# Column number containing QPN
MFG_col 	= 0						# Column location for manufacturer part number
DES_col 	= 0 					# Column location for descritpion part number
QTY_col 	= 0 					# Column location for quantity field
CR1_col		= 0						# Column location for supplier name
CR1PN_col	= 0						# Column location for supplier's PN
NOTE_col 	= 0 					# Column location for "notes" field
BOM_HEADER 	= ["QPN","QTY","DES","MFG","MFGPN","CR1","CR1PN","NOTES"]
EXTS		= ('.xls','.xlsx')		#Supported file extensions

MAX_HITS = 5			#This many hits will trigger us to leave the searching loop
data_start = 0			#This is the row where the data starts

search_header 	= []		# Set equal to BOM_HEADER and pop elements until we find all the colums we're looking for
header 			= []		# This array will define the column locations for the header
qpn 			= []        # Pull in all QPNs into a list. This will make them easier to work with later
asso 			= []       	# Pull in all associations into a list. This will make them easier to work with later
qty 			= []        # Pull in all QTYs into a list. This will make them easier to work with later
des 			= []		# Pull in all Descriptions into a list. This will make them easier to work with later
mfg 			= []		# Pull in all Manufactures into a list. This will make them easier to work with later
mfgpn 			= []		# Pull in all Manufacturing Part Numbers into a list. This will make them easier to work with later
cr1 			= []		# Pull in all suppler names into a list. This will make them easier to work with later
cr1pn 			= []		# Pull in all supplier pn's into a list. This will make them easier to work with later
notes 			= []		# Pull all note values into a list. This will make them easier to work with later


## FUNCTIONS ##
######################	   
def debugbreak():
	while(1):
		pass
		
def clean_value(textin):
	temptext = textin
	temptext = temptext.lstrip('text:u\'')     	# Remove the initial part of the string that we don't need 'text:u'   
	temptext = temptext.replace("'","")			# Remove single quote marks from value
	temptext = temptext.strip()					# Remove only leading and trailing white spaces
	if(temptext.find("number:") != -1):
		temptext = temptext.replace("number:","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
	
def clean_des(textin):
	temptext = textin
	temptext = temptext.lstrip('text:u\'')     #Remove the initial part of the string that we don't need 'text:u'  
	temptext = temptext.replace("'","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
def pause():
	user_input=raw_input("Press any key to exit...")
	sys.exit(0)
	   
#****************************************************************************** 
#******************************  ---MAIN---  **********************************
#******************************************************************************   
if __name__ == '__main__':

	path = os.getcwd()
	# Find path/dirs/files
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files

	print "Files found in directory: " + str(len(files))

	# Iterate through all files in directory
	for i in range(len(files)-1):
		
		# Search through and open files that are appended with *.xlsx
		if(files[i].lower().endswith(ext) for ext in EXTS):
			
			print "Opening file: ", files[i]
			wb = load_workbook(filename = files[i])     #Open the workbook that we are going to parse though 
			ws = wb.sheetnames             #Grab the names of the worksheets -- I believe this line is critical.
			
			num_sheets = len(ws)				#This is the number of sheet

			print "\n\n==============================================="
			print "==============================================="
			print "File opened: " + str(files[i])
			print "The number of worksheets is: " + str(num_sheets)
			print "+++++++++++++++++++++++++++++++++++++++++++++++"
			
			# Iterate through all sheets in the opened file
			for sh in range (num_sheets):
				current_sheet = wb.get_sheet_by_name(name = ws[sh])
				association = raw_input("Enter a unique association / high-level QPN for this worksheet (i.e. Prog Cbl): ") 
				
				num_rows = current_sheet.max_row      #This is how many rows are in the worksheet
				num_cols = current_sheet.max_column 		#This is how many columns are in the sheet

				for r in range (1,num_rows + 1):			# Find the header locations. Excel starts counting at one
					search_header = ["QPN","QTY","DES","MFG","MFGPN","CR1","CR1PN","NOTES"]					# Load up headers we need to search for
					print 'Search header before starting: ', search_header

					# Iterate over columns of current row
					for c in range (1,num_cols + 1):				# Excel starts counting at 1
						temptext = unicode(current_sheet.cell(row = r, column=c).value)                  #This is the fifth cell of the current row
						temptext = temptext.lstrip('text:u\'')     #Remove the initial part of the string that we don't need 'text:u'   
						temptext = temptext.rstrip('\'')     		#Remove the initial part of the string that we don't need 'text:u'   
						temptext = temptext.replace(" ","")			#This will remove any and all white spaces

						
						if((temptext.find("QPN")!=-1) or (temptext.find("qpn")!=-1)):
							QPN_col = c
							search_header.remove("QPN")
						
						elif((temptext.find("MFGPN")!=-1) or (temptext.find("MFG PN")!=-1)):	#Look for MFGPN header
							MFGPN_col = c
							search_header.remove("MFGPN")
						
						elif((temptext.find("MFG")!=-1) and (temptext.find("PN") == -1)):		#Look for MFG -- make sure PN is not present
							MFG_col = c
							search_header.remove("MFG")
						
						elif((temptext.find("Des")!=-1) or (temptext.find("DES")!=-1) or (temptext.find("Description")!=-1) or (temptext.find("DESCRIPTION")!=-1)):		#Look for Description
							DES_col = c
							search_header.remove("DES")
						
						elif((temptext.find("Qty")!=-1) or (temptext.find("QTY")!=-1)):		#Look for Quantity field.  
							search_header.remove("QTY")
							QTY_col = c
						
						elif( ((temptext.find("cr1")!=-1) or (temptext.find("CR1")!=-1)) and (len(temptext) <= 4) ):		#Look for CR1, and cannot have PN as in CR1PN
							search_header.remove("CR1")
							CR1_col = c
						
						elif( ((temptext.find("cr1pn")!=-1) or (temptext.find("CR1PN")!=-1)) and (len(temptext) > 4) ):		#Look for CR1PN
							search_header.remove("CR1PN")
							CR1PN_col = c
						
						elif((temptext.find("notes")!=-1) or (temptext.find("NOTES")!=-1) or (temptext.find("Notes")!=-1) or (temptext.find("Note")!=-1) or (temptext.find("note")!=-1) ):		#Look for Notes 
							search_header.remove("NOTES")
							NOTE_col = c
					
					if( (len(search_header) == 0) ):		# Found all header fields
						data_start = r + 1			# Plenty of confidence at this point that we've found data start
						print 'Data appears to start on row: ', data_start
						
						print( 	'Sample data in start row: ', clean_value(unicode(current_sheet.cell(row = data_start, column=QPN_col).value)),' ', 
								clean_value(unicode(current_sheet.cell(row = data_start, column=DES_col).value)), ' ', 
								clean_value(unicode(current_sheet.cell(row = data_start, column=MFG_col).value)), ' ', 
								clean_value(unicode(current_sheet.cell(row = data_start, column=MFGPN_col).value))
							)
						break
					
					elif( (r == 10) and (len(search_header) > 0) ):		# Some header fields are missing, so shutdown
						print "On sheet: " + str(current_sheet) + " -- did not find headers: ", search_header
						sys.exit(0)
				
				print "QPN column found to be: " + str(QPN_col)		
				print "QTY column found to be: " + str(QTY_col)
				print "Description column found to be: " + str(DES_col)		
				print "MFG column found to be: " + str(MFG_col)
				print "MFGPN column found to be: " + str(MFGPN_col)
				print "CR1 column found to be: " + str(CR1_col)
				print "CR1PN column found to be: " + str(CR1PN_col)
				print "NOTES column found to be: " + str(NOTE_col)
				
				header = [0,QPN_col,DES_col,MFG_col,MFGPN_col,CR1_col,CR1PN_col,QTY_col,NOTE_col]
				header_values = ["Association","QPN","DES","MFG","MFGPN","CR1","CR1PN","QTY","NOTES"]
				
				# Now iterate through all rows of the current sheet and populate the data lists
				blank_row_count = 0		# Reset number of blank rows detected.  When three in a row are detected, break out of the loop. 
				for r in range (data_start,num_rows + 1):
					
					
					# If multiple columns are blank, break out of this loop for these are empty cells
					if( (len(clean_value(unicode(current_sheet.cell(row = r, column=QPN_col).value))) <= 1) and ( len(clean_des(unicode(current_sheet.cell(row = r, column=DES_col).value))) <= 1) and
						( len(clean_value(unicode(current_sheet.cell(row = r, column=MFG_col).value))) <= 1) ):
						blank_row_count += 1				# Increase value of blank row count
						print 'Blank row detected at row (', r, ')'
					else:
						blank_row_count = 0					
						asso.append(association)				#For each row in the BOM, we need to append the association
						print( 	'Sample data, current row: ', clean_value(unicode(current_sheet.cell(row = r, column=QPN_col).value)),' ', 
								clean_value(unicode(current_sheet.cell(row = r, column=DES_col).value)), ' ', 
								clean_value(unicode(current_sheet.cell(row = r, column=MFG_col).value)), ' ', 
								clean_value(unicode(current_sheet.cell(row = r, column=MFGPN_col).value))
							)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=QPN_col).value))
						if current_value == "None":
							current_value = ""
						qpn.append(current_value)			
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=DES_col).value))
						if current_value == "None":
							current_value = ""
						des.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=MFG_col).value))
						if current_value == "None":
							current_value = ""
						mfg.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=MFGPN_col).value))
						if current_value == "None":
							current_value = ""
						mfgpn.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=CR1_col).value))
						if current_value == "None":
							current_value = ""
						cr1.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=CR1PN_col).value))
						if current_value == "None":
							current_value = ""
						cr1pn.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=QTY_col).value))
						if current_value == "None":
							current_value = ""
						qty.append(current_value)
						
						current_value = clean_value(unicode(current_sheet.cell(row = r, column=NOTE_col).value))
						if current_value == "None":
							current_value = ""
						notes.append(current_value)

					if(blank_row_count >= 3):
						break								# Too many blank rows detected, so break out of the loop.  
					

	print "\n+++++++++++++++++++++++++++++++++++++++++++++++"
	print "+++++++++++++++++++++++++++++++++++++++++++++++"
	print "Creating combined BOM"
	
	NewBook = Workbook()
	NewSheet = NewBook.active
	NewSheet.title = "Combined BOM"

	# Write the header values
	for i in range (1,len(header)+1):
		NewSheet.cell(row=1,column=i).value = header_values[i-1]
	
	# Write rows of the combined BOM
	for i in range (2,len(asso) + 2):				
		NewSheet.cell(row=i,column=1).value = asso[i-2]
		NewSheet.cell(row=i,column=2).value = qpn[i-2]
		NewSheet.cell(row=i,column=3).value = des[i-2]
		NewSheet.cell(row=i,column=4).value = mfg[i-2]
		NewSheet.cell(row=i,column=5).value = mfgpn[i-2]
		NewSheet.cell(row=i,column=6).value = cr1[i-2]
		NewSheet.cell(row=i,column=7).value = cr1pn[i-2]
		NewSheet.cell(row=i,column=8).value = qty[i-2]
		NewSheet.cell(row=i,column=9).value = notes[i-2]
		print ".",

	NewBook.save(filename = "CombinedBOM.xlsx")
	print " "
	null=raw_input("Press any key to close...")