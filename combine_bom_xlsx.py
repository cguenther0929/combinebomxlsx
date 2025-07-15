"""
FILE: Combined_BOM.py

PURPOSE: 
When running this script, the user will be able to create one large flat BOM
for purchasing purposes.  This script will only operate on .xlsx files, and not
.xls files.  This script will automatically sift through all files in the 
current working directory, and with each file, it will iterate over all sheets. 

AUTHOR: 
Clinton G. 

TODO: Nothing

"""
import sys
import random
import time
import csv
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

# ----------------------------------------------------------------------- #
# Regular Expression Strings
# ----------------------------------------------------------------------- #
qpn_re 		= "(QPN)"
mfgpn_re 	= "(MFG.?PN)"						# To match MFGPN or MFG PN (will ignore case)
mfg_re 		= "(MFG)|(MANUFACTURER)"			
des_re 		= "(DES)|(DESCRIPTION)"
ref_re		= "(REF)|(REF.DES)|(REFERENCE)"
qty_re		= "(QTY)|(QUANTITY)"
uom_re		= "(UOM)|(UNIT OF MEASURE)"
cr1_re		= "(CR1)"
cr1pn_re	= "(CR1PN)"
notes_re	= "(NOTES)|(NOTE)"



## DEFINE VRIABLES ##
#####################
MFGPN_col 	= 0						# Column number containing the MFGPN
QPN_col 	= 0						# Column number containing QPN
MFG_col 	= 0						# Column location for manufacturer part number
DES_col 	= 0 					# Column location for description part number
QTY_col 	= 0 					# Column location for quantity field
UOM_col 	= 0 					# Column location for UOM field
CR1_col		= 0						# Column location for supplier name
CR1PN_col	= 0						# Column location for supplier's PN
NOTE_col 	= 0 					# Column location for "notes" field
BOM_HEADER 	= ["QPN","QTY","UOM","DES","REF","MFG","MFGPN","CR1","CR1PN","NOTES"]

data_start 				= 0			# This is the row where the data starts
flag_header_detected 	= False		# Set to true as soon as we detect header data in one of the rows

search_header 	= []		# Set equal to BOM_HEADER and pop elements until we find all the colums we're looking for
sheet_valid		= False		# Flag that tells application if a sheet contains valid data or not
header 			= []		# This array will define the column locations for the header
qpn 			= []        # Pull in all QPNs into a list. This will make them easier to work with later
asso 			= []       	# Pull in all associations into a list. This will make them easier to work with later
qty 			= []        # Pull in all QTYs into a list. This will make them easier to work with later
uom 			= []        # Pull in all UOM values into a list. This will make them easier to work with later
des 			= []		# Pull in all Descriptions into a list. This will make them easier to work with later
ref 			= []		# Pull all reference values into a list. This will make them easier to work with later
mfg 			= []		# Pull in all Manufactures into a list. This will make them easier to work with later
mfgpn 			= []		# Pull in all Manufacturing Part Numbers into a list. This will make them easier to work with later
cr1 			= []		# Pull in all suppler names into a list. This will make them easier to work with later
cr1pn 			= []		# Pull in all supplier pn's into a list. This will make them easier to work with later
notes 			= []		# Pull all note values into a list. This will make them easier to work with later


# -------------------------------------- #
# Local Methods
# -------------------------------------- #
def debugbreak():
	while(1):
		pass
		
def clean_value(textin):
	temptext = textin
	# logging.info("Text entered into method clean value: " + str(temptext))
	temptext = temptext.lstrip('text:u\'')     	# Remove the initial part of the string that we don't need 'text:u'   
	temptext = temptext.lstrip("b\'")     	# Remove the initial part of the string that we don't need 'text:u'   
	temptext = temptext.replace("'","")			# Remove single quote marks from value
	temptext = temptext.strip()					# Remove only leading and trailing white spaces
	if(temptext.find("number:") != -1):
		temptext = temptext.replace("number:","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
	
def clean_des(textin):
	temptext = textin
	temptext = temptext.lstrip('text:u\'')     #Remove the initial part of the string that we don't need 'text:u'  
	temptext = temptext.replace("'","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
def pause():
	user_input=input("Press any key to exit...")
	sys.exit(0)

# -------------------------------------- #
# Setup Logging
# -------------------------------------- #
logging.basicConfig(
    filename = 'combine_bom.log',
    level = logging.DEBUG,
    format =' %(asctime)s -  %(levelname)s - %(message)s',
	filemode = 'w'
)
	   
#****************************************************************************** 
#******************************  ---MAIN---  **********************************
#******************************************************************************   
if __name__ == '__main__':

	
	# ----------------------------------------------------------------------- #
	# Build the directory tree
	# ----------------------------------------------------------------------- #
	path = os.getcwd()
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files
	
	# ----------------------------------------------------------------------- #
	# If a consolidated BOM already exists, then remove it
	# ----------------------------------------------------------------------- #
	for i in range(len(files)):
		if(files[i].find("Combined") != -1):
			logging.info("Deleting existing consolidated BOM.")
			os.remove(files[i])
	
	# ----------------------------------------------------------------------- #
	# Some file may have been removed, so refresh 
	# directory information.  
	# ----------------------------------------------------------------------- #
	path = os.getcwd()
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files

	print ("Files found in directory: ", str(len(files)))
	print ("File names: ", files)

	# ----------------------------------------------------------------------- #
	# Iterate through files
	# ----------------------------------------------------------------------- #
	for i in range(len(files)):
		# ----------------------------------------------------------------------- #
		# Search through files and open only those having the proper extension 
		# ----------------------------------------------------------------------- #
		if(files[i].upper().endswith(".XLSX")):
			
			print ("Opening file: ", files[i])
			wb = load_workbook(filename = files[i])     # Open the workbook that we are going to parse though 
			ws = wb.sheetnames             				# Grab the names of the worksheets -- I believe this line is critical.
			
			num_sheets = len(ws)				#This is the number of sheet

			print ("\n\n===============================================")
			print ("===============================================")
			print ("File opened: ", str(files[i]))
			print ("The number of worksheets is: ", str(num_sheets))
			print ("Worksheet names: ", ws)
			print ("+++++++++++++++++++++++++++++++++++++++++++++++")
			
			logging.info ("===============================================")
			logging.info ("===============================================")
			logging.info ("File opened: " + str(files[i]))
			logging.info ("The number of worksheets is: " + str(num_sheets))
			logging.info ("Worksheet names: " + str(ws))
			logging.info ("+++++++++++++++++++++++++++++++++++++++++++++++")
			
			# ----------------------------------------------------------------------- #
			# Iterate through all sheets
			# ----------------------------------------------------------------------- #
			for sh in range (num_sheets):
				sheet_valid = False
				
				current_sheet = wb[ws[sh]]
				
				print("Now operating on worksheet: ", ws[sh])
				association = input("Enter a unique association / high-level QPN for this worksheet (i.e. Prog Cbl): ") 
				
				num_rows = current_sheet.max_row     		
				num_cols = current_sheet.max_column 		

				# ----------------------------------------------------------------------- #
				# Iterate through every row on current sheet
				# ----------------------------------------------------------------------- #
				for r in range (1,num_rows + 1):					# Find the header locations. Excel starts counting at one
					search_header = BOM_HEADER.copy()						# Load up headers we need to search for
					print ("Search header before starting: ", search_header)
					
					flag_header_detected = False
					# ----------------------------------------------------------------------- #
					# Iterate over columns of selected row
					# ----------------------------------------------------------------------- #
					for c in range (1,num_cols + 1):				# Excel starts counting at 1
						
						temptext = str(str(current_sheet.cell(row = r, column=c).value).encode(encoding = 'UTF-8',errors = 'strict'))                
						temptext = temptext.lstrip("text:u\'")     	# Clean up the extra garbage on text 
						temptext = temptext.lstrip("b\'")     		
						temptext = temptext.rstrip("\'")     		
						temptext = temptext.replace(" ","")			# Remove any and all white spaces 
						# logging.info("Text extracted from cell: " + temptext)
						# print ("****DEBUG Text Extracted: ", temptext)
						# print ("****DEBUG Current column number: ", str(c))


						
						if(re.fullmatch(qpn_re,temptext,re.IGNORECASE)):
							QPN_col = c
							search_header.remove("QPN")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found QPN")
						
						elif(re.fullmatch(mfgpn_re,temptext,re.IGNORECASE)):	#Look for MFGPN header
							MFGPN_col = c
							search_header.remove("MFGPN")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found MFGPN")
						
						elif(re.fullmatch(mfg_re,temptext,re.IGNORECASE)):		#Look for MFG -- make sure PN is not present
							MFG_col = c
							search_header.remove("MFG")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found MFG")
						
						elif(re.fullmatch(des_re,temptext,re.IGNORECASE)):		#Look for Description
							DES_col = c
							search_header.remove("DES")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found DES")
						
						elif(re.fullmatch(ref_re,temptext,re.IGNORECASE)):		#Look for Description
							REF_col = c
							search_header.remove("REF")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found REF")
						
						elif(re.fullmatch(qty_re,temptext,re.IGNORECASE)):		#Look for Quantity field.  
							QTY_col = c
							search_header.remove("QTY")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found QTY")

						elif(re.fullmatch(uom_re,temptext,re.IGNORECASE)):		#Look for Quantity field.  
							UOM_col = c
							search_header.remove("UOM")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found UOM")
						
						elif(re.fullmatch(cr1_re,temptext,re.IGNORECASE)):		#Look for CR1, and cannot have PN as in CR1PN
							CR1_col = c
							search_header.remove("CR1")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found CR1")
						
						elif(re.fullmatch(cr1pn_re,temptext,re.IGNORECASE)):		#Look for CR1PN
							CR1PN_col = c
							search_header.remove("CR1PN")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found CR1PN")
						
						elif(re.fullmatch(notes_re,temptext,re.IGNORECASE)):		#Look for Notes 
							NOTE_col = c
							search_header.remove("NOTES")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found NOTES")

						# High confidence that we've found the header
						if(len(search_header) <= 2):
							flag_header_detected = True
							data_start = r + 1			# data should start on following row
					
					if( (len(search_header) == 0) ):		# Found all header fields
						sheet_valid = True
						print ("Data appears to start on row: ", data_start)
						
						print( 	"Sample data in start row: ", clean_value(str(str(current_sheet.cell(row = data_start, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict'))),' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=MFG_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=MFGPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							)
						break

					elif((len(search_header) == 1)):		# This BOM does not contain the reference field
						
						reference_index = 0
						try:
							reference_index = search_header.index("REF")
						except:
							pass
					
						if(reference_index > 0):
							REF_col = 1
							# search_header.remove("REF")
							sheet_valid = True
							print ("This BOM does not appear to contain a reference column.")
							logging.info ("This BOM does not appear to contain a reference column.")
							print ("Data appears to start on row: ", data_start)
							logging.info("Data appears to start on row: " + str(data_start))
							
							print( 	"Sample data in start row: ", clean_value(str(str(current_sheet.cell(row = data_start, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict'))),' ', 
									clean_value(str(str(current_sheet.cell(row = data_start, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = data_start, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = data_start, column=MFG_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = data_start, column=MFGPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
								)
							break
					
					elif(flag_header_detected == True):
						sheet_valid = False
						print ("Found valid header data, but missing: ", search_header)
						logging.info("Found valid header data, but missing: " + str(search_header))
						logging.info("System existing due to an invalid BOM format.")
						
						print("\n\n====================================")
						print("Exiting due to invalid BOM format.")
						userinput = input("Press any key to exit...")
						sys.exit(0)

				if(sheet_valid):
					print ("QPN column found to be: ", 			str(QPN_col))		
					print ("QTY column found to be: ", 			str(QTY_col))
					print ("UOM column found to be: ", 			str(UOM_col))
					print ("Description column found to be: ", 	str(DES_col))		
					print ("Reference column found to be: ", 	str(REF_col))		
					print ("MFG column found to be: ", 			str(MFG_col))
					print ("MFGPN column found to be: ", 		str(MFGPN_col))
					print ("CR1 column found to be: ", 			str(CR1_col))
					print ("CR1PN column found to be: ", 		str(CR1PN_col))
					print ("NOTES column found to be: ", 		str(NOTE_col))
					
					header = [0,QPN_col,DES_col,REF_col,MFG_col,MFGPN_col,CR1_col,CR1PN_col,QTY_col,UOM_col,NOTE_col]
					header_values = ["Association","QPN","DES","REF","MFG","MFGPN","CR1","CR1PN","QTY","UOM","NOTES"]
					
					# Now iterate through all rows of the current sheet and populate the data lists
					blank_row_count = 0		# Reset number of blank rows detected.  When three in a row are detected, break out of the loop. 
					for r in range (data_start,num_rows + 1):
						
						
						# If multiple columns are blank, break out of this loop for these are empty cells
						if( ( len(clean_des(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) and
						 	( len(clean_des(str(str(current_sheet.cell(row = r, column=MFG_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) and
							( len(clean_des(str(str(current_sheet.cell(row = r, column=MFGPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) ):
							
							blank_row_count += 1				# Increase value of blank row count
							print ("Blank row detected at row (", r, ")")
						
						else:
							
							blank_row_count = 0					
							asso.append(association)				# For each row in the BOM, we need to append the association
							print( 	'Sample data, current row: ', 
									clean_value(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = r, column=MFG_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = r, column=MFGPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
								)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							qpn.append(current_value)			
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							des.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							ref.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=MFG_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							mfg.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=MFGPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							mfgpn.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=CR1_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							cr1.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=CR1PN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							cr1pn.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=QTY_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							qty.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=UOM_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							uom.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=NOTE_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							notes.append(current_value)

						if(blank_row_count >= 3):
							break								# Too many blank rows detected, so break out of the loop.  
					

	print ("\n+++++++++++++++++++++++++++++++++++++++++++++++")
	print ("+++++++++++++++++++++++++++++++++++++++++++++++")
	print ("Creating combined BOM")
	
	NewBook = Workbook()
	NewSheet = NewBook.active
	NewSheet.title = "Combined BOM"

	# Write the header values
	for i in range (1,len(header)+1):
		NewSheet.cell(row=1,column=i).value = header_values[i-1]
	
	# Write rows of the combined BOM
	for i in range (2,len(asso) + 2):				
		NewSheet.cell(row=i,column=1).value = asso[i-2]
		NewSheet.cell(row=i,column=2).value = qpn[i-2]
		NewSheet.cell(row=i,column=3).value = des[i-2]
		NewSheet.cell(row=i,column=4).value = ref[i-2]
		NewSheet.cell(row=i,column=5).value = mfg[i-2]
		NewSheet.cell(row=i,column=6).value = mfgpn[i-2]
		NewSheet.cell(row=i,column=7).value = cr1[i-2]
		NewSheet.cell(row=i,column=8).value = cr1pn[i-2]

		NewSheet.cell(row=i,column=9).number_format = '0.00'
		NewSheet.cell(row=i,column=9).value = qty[i-2]

		NewSheet.cell(row=i,column=10).value = uom[i-2]

		NewSheet.cell(row=i,column=11).value = notes[i-2]
		print (".", end = ' '),

	NewBook.save(filename = "CombinedBOM.xlsx")
	print (" ")
	null=input("Press any key to close...")